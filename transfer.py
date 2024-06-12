import csv
from openpyxl import Workbook
from openpyxl.styles import Font

rows=[]
def get_teachers():
    with open('allotment.csv', 'r') as csvfile:
        csvreader = csv.reader(csvfile)
        for row in csvreader:
            rows.append(row)
    teachers=[]
    offset=0
    for i in rows:
        dict={'teacher':i[1]}
        if dict not in teachers:
            teachers.append(dict)
    for i in teachers:
        i['offset']=offset
        offset+=10
    print(teachers)
    print("5: End")
    return teachers

def teacher_transfer():
    print("6: Start")
    wb = Workbook()
    days=['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    teachers =get_teachers()
    ws = wb.active
    offset=0

    for j in range(len(teachers)): 
        for i in range(2, 9):
            ws.cell(row=1+offset,column=i).value=i-1
            ws.cell(row=1+offset,column=i).font=Font(bold=True)
        for i in range(2,8):
            ws.cell(row=i+offset,column=1).value=days[i-2]
            ws.cell(row=i+offset,column=1).font=Font(bold=True)
        offset+=10

    for j in teachers: 
        ws.cell(row=1+j['offset'],column=1).value=j['teacher']
        ws.cell(row=1+j['offset'],column=1).font=Font(bold=True)

    for period in rows:
        teacher=period[1]
        offset=0
        for i in teachers:
            if i['teacher']==teacher:
                offset=i['offset']
                break
        day=days.index(period[6])+1
        hour=int(period[5])
        temp=ws.cell(row=day+offset+1,column=hour+1).value
        if temp:
            ws.cell(row=day+offset+1,column=hour+1).value =period[2]+" ("+str(period[3])+" "+period[4]+") "+temp
        else:
            ws.cell(row=day+offset+1,column=hour+1).value =period[2]+" ("+str(period[3])+" "+period[4]+")"

    wb.save('teacher_time_tables.xlsx')